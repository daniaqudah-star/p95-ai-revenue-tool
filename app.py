Yes — delete everything in `app.py` and paste this full version.

This version uses:

```text
Budget lines from: Workload and Resources P1
Row start date from: start time column
Row end date from: end time column
Fallback timeline: project-level timeline if row date is missing
Revenue spread: based on each row’s own timeline + phase/unit logic
```

```python
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta
from collections import defaultdict
from io import BytesIO

st.set_page_config(page_title="P95 AI Revenue Module Generator", layout="wide")

st.title("P95 AI Revenue Module Generator")
st.caption("PMO Internal Tool • Clinical Study Revenue Automation")


def month_key(dt):
    return datetime(dt.year, dt.month, 1)


def safe_text(value):
    if pd.isna(value):
        return ""
    return str(value).strip()


def to_datetime_or_none(value):
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return None
    if parsed.year < 2020 or parsed.year > 2035:
        return None
    return parsed.to_pydatetime()


def get_months_between(start, end):
    months = []
    current = datetime(start.year, start.month, 1)
    final = datetime(end.year, end.month, 1)

    while current <= final:
        months.append(current)
        if current.month == 12:
            current = datetime(current.year + 1, 1, 1)
        else:
            current = datetime(current.year, current.month + 1, 1)

    return months


def find_study_timeline_from_all_sheets(budget_file):
    all_sheets = pd.read_excel(budget_file, sheet_name=None, header=None)
    date_pairs = []

    for sheet_name, df in all_sheets.items():
        for row_idx in range(len(df)):
            dates = []

            for value in df.iloc[row_idx].tolist():
                parsed = to_datetime_or_none(value)

                if parsed:
                    dates.append(parsed)

            if len(dates) >= 2:
                start = min(dates)
                end = max(dates)
                duration_days = (end - start).days

                if 14 <= duration_days <= 3000:
                    date_pairs.append({
                        "sheet": sheet_name,
                        "row": row_idx + 1,
                        "start": start,
                        "end": end,
                        "duration_days": duration_days
                    })

    if not date_pairs:
        raise ValueError("Could not detect project-level timeline from any budget sheet.")

    date_pairs = sorted(date_pairs, key=lambda x: x["duration_days"], reverse=True)
    best = date_pairs[0]

    return best["start"], best["end"], best


def find_timeline_from_workload_sheet(df):
    date_pairs = []

    for row_idx in range(len(df)):
        dates = []

        for value in df.iloc[row_idx].tolist():
            parsed = to_datetime_or_none(value)

            if parsed:
                dates.append(parsed)

        if len(dates) >= 2:
            start = min(dates)
            end = max(dates)
            duration_days = (end - start).days

            if 14 <= duration_days <= 3000:
                date_pairs.append({
                    "sheet": "Workload and Resources P1",
                    "row": row_idx + 1,
                    "start": start,
                    "end": end,
                    "duration_days": duration_days
                })

    if date_pairs:
        date_pairs = sorted(date_pairs, key=lambda x: x["duration_days"], reverse=True)
        best = date_pairs[0]
        return best["start"], best["end"], best

    return None, None, None


def detect_header_row(df):
    keywords = [
        "activities",
        "activity",
        "unit id",
        "units",
        "# of units",
        "unit cost",
        "total direct costs",
        "start time",
        "end time"
    ]

    for idx in range(min(len(df), 100)):
        row_values = [safe_text(x).lower() for x in df.iloc[idx].tolist()]
        row_text = " ".join(row_values)

        matches = sum(1 for keyword in keywords if keyword in row_text)

        if matches >= 3:
            return idx

    return 0


def extract_budget_activities(budget_file):
    sheet_name = "Workload and Resources P1"

    df = pd.read_excel(
        budget_file,
        sheet_name=sheet_name,
        header=None
    )

    header_row = detect_header_row(df)

    headers = [
        safe_text(x).lower()
        for x in df.iloc[header_row].tolist()
    ]

    def find_col(possible_names, fallback=None):
        for name in possible_names:
            for i, header in enumerate(headers):
                if name in header:
                    return i
        return fallback

    activity_col = find_col(["activities", "activity"], 1)
    description_col = find_col(["unit id", "unit description", "description"], 2)
    units_col = find_col(["# of units", "units"], 4)
    unit_price_col = find_col(["unit cost", "unit price"], 6)
    total_price_col = find_col(["total direct costs", "total price", "total cost"], 7)
    row_start_col = find_col(["start time", "start date"], 8)
    row_end_col = find_col(["end time", "end date"], 9)

    activities = []

    for idx in range(header_row + 1, len(df)):
        row = df.iloc[idx]

        activity = row[activity_col] if activity_col is not None and activity_col < len(row) else None
        description = row[description_col] if description_col is not None and description_col < len(row) else ""
        units = row[units_col] if units_col is not None and units_col < len(row) else None
        unit_price = row[unit_price_col] if unit_price_col is not None and unit_price_col < len(row) else None
        total_price = row[total_price_col] if total_price_col is not None and total_price_col < len(row) else None

        row_start_raw = row[row_start_col] if row_start_col is not None and row_start_col < len(row) else None
        row_end_raw = row[row_end_col] if row_end_col is not None and row_end_col < len(row) else None

        row_start = to_datetime_or_none(row_start_raw)
        row_end = to_datetime_or_none(row_end_raw)

        activity_text = safe_text(activity)
        description_text = safe_text(description)

        if not activity_text:
            continue

        lower_text = activity_text.lower()

        skip_terms = [
            "insert lines",
            "sectiontotal",
            "section total",
            "budget total",
            "project budget",
            "assumptions",
            "timelines",
            "meetings",
            "rates",
            "resource",
            "inflation",
            "total",
            "name"
        ]

        if any(term == lower_text or lower_text.startswith(term) for term in skip_terms):
            continue

        units_numeric = pd.to_numeric(units, errors="coerce")
        unit_price_numeric = pd.to_numeric(unit_price, errors="coerce")
        total_numeric = pd.to_numeric(total_price, errors="coerce")

        if pd.isna(units_numeric):
            units_numeric = 1

        if pd.isna(unit_price_numeric):
            if not pd.isna(total_numeric) and float(units_numeric) != 0:
                unit_price_numeric = float(total_numeric) / float(units_numeric)
            else:
                continue

        if float(units_numeric) == 0 or float(unit_price_numeric) == 0:
            continue

        activities.append({
            "activity": activity_text,
            "description": description_text,
            "units": float(units_numeric),
            "unit_price": float(unit_price_numeric),
            "total_price": total_numeric,
            "row_start": row_start,
            "row_end": row_end,
            "source_sheet": sheet_name,
            "source_row": idx + 1
        })

    if not activities:
        raise ValueError(
            "Could not detect usable budget activity rows from 'Workload and Resources P1'."
        )

    return activities, sheet_name, header_row + 1, df


def get_phase_dates(start, end):
    total_days = max((end - start).days, 1)

    startup_end = start + timedelta(days=int(total_days * 0.25))
    execution_end = start + timedelta(days=int(total_days * 0.75))
    analysis_end = start + timedelta(days=int(total_days * 0.90))

    return {
        "startup": (start, startup_end),
        "execution": (startup_end, execution_end),
        "analysis": (execution_end, analysis_end),
        "closeout": (analysis_end, end),
    }


def assign_phase(activity, description):
    text = f"{activity} {description}".lower()

    if any(x in text for x in [
        "contract signature",
        "protocol",
        "sample size",
        "kom",
        "kick off",
        "kick-off",
        "training",
        "development",
        "submission",
        "approval",
        "start-up",
        "startup",
        "set-up",
        "setup"
    ]):
        return "startup"

    if any(x in text for x in [
        "meeting",
        "monthly",
        "bi-weekly",
        "biweekly",
        "tc",
        "coordination",
        "internal",
        "data collection",
        "monitoring",
        "management",
        "operational",
        "site",
        "country",
        "ctms",
        "tmf"
    ]):
        return "execution"

    if any(x in text for x in [
        "review",
        "analysis",
        "database lock",
        "stat",
        "biostat",
        "data review"
    ]):
        return "analysis"

    if any(x in text for x in [
        "close-out",
        "closeout",
        "closure",
        "archive",
        "archiving",
        "transfer",
        "final"
    ]):
        return "closeout"

    return "execution"


def spread_units(activity, description, total_units, start, end):
    phase_dates = get_phase_dates(start, end)
    phase = assign_phase(activity, description)

    phase_start, phase_end = phase_dates[phase]

    # If the row already has its own specific timeline,
    # we spread within that row timeline but still use activity type.
    phase_months = get_months_between(phase_start, phase_end)

    spread = defaultdict(float)
    total_units = float(total_units)

    combined_text = f"{activity} {description}".lower()

    if len(phase_months) == 0:
        spread[month_key(start)] = total_units
        return spread

    if any(x in combined_text for x in [
        "monthly",
        "months",
        "bi-weekly",
        "biweekly",
        "weekly",
        "meeting",
        "meetings",
        "tc",
        "coordination",
        "management",
        "updates"
    ]):
        per_month = total_units / len(phase_months)
        for m in phase_months:
            spread[month_key(m)] = per_month

    elif any(x in combined_text for x in [
        "document",
        "process",
        "contract signature",
        "kom",
        "training",
        "set-up",
        "setup"
    ]):
        spread[month_key(phase_start)] = total_units

    elif any(x in combined_text for x in [
        "review",
        "round",
        "analysis"
    ]):
        per_month = total_units / len(phase_months)
        for m in phase_months:
            spread[month_key(m)] = per_month

    else:
        spread[month_key(phase_start)] = total_units

    return spread


def generate_revenue_tracker(budget_file, template_file):
    activities, activity_sheet, header_row, workload_df = extract_budget_activities(budget_file)

    project_start, project_end, timeline_source = find_timeline_from_workload_sheet(workload_df)

    if project_start is None or project_end is None:
        project_start, project_end, timeline_source = find_study_timeline_from_all_sheets(budget_file)

    wb = load_workbook(
        template_file,
        keep_links=True,
        data_only=False
    )

    if "UNIT TRACKER" not in wb.sheetnames:
        raise ValueError("UNIT TRACKER sheet not found in template.")

    ws = wb["UNIT TRACKER"]

    ws["G5"] = project_start
    ws["G6"] = project_end

    project_months = get_months_between(project_start, project_end)

    forecast_cols = {}
    start_col = 18  # R

    for i, m in enumerate(project_months):
        forecast_cols[month_key(m)] = start_col + (i * 6)

    target_row = 40
    missing_row_timeline_count = 0

    for item in activities:
        activity = item["activity"]
        description = item["description"]
        units = item["units"]
        unit_price = item["unit_price"]

        row_start = item["row_start"] or project_start
        row_end = item["row_end"] or project_end

        if item["row_start"] is None or item["row_end"] is None:
            missing_row_timeline_count += 1

        if row_end < row_start:
            row_start, row_end = row_end, row_start

        ws[f"D{target_row}"] = row_start
        ws[f"E{target_row}"] = row_end
        ws[f"F{target_row}"] = activity
        ws[f"H{target_row}"] = description
        ws[f"I{target_row}"] = units
        ws[f"J{target_row}"] = unit_price

        unit_spread = spread_units(
            activity,
            description,
            units,
            row_start,
            row_end
        )

        total_forecast_units = 0

        for m, forecast_col in forecast_cols.items():
            forecast_units = unit_spread.get(m, 0)
            ws.cell(row=target_row, column=forecast_col).value = forecast_units
            total_forecast_units += forecast_units

        ws[f"M{target_row}"] = total_forecast_units

        target_row += 1

    validation_warnings = []
    rows_processed = 0
    total_contract_value = 0
    total_forecast_units = 0

    for row in range(40, target_row):
        activity = ws[f"F{row}"].value
        unit_description = ws[f"H{row}"].value
        contracted_units = ws[f"I{row}"].value
        unit_cost = ws[f"J{row}"].value
        forecast_units = ws[f"M{row}"].value

        rows_processed += 1

        if not activity:
            validation_warnings.append(f"Row {row}: Missing activity name")

        if not unit_description:
            validation_warnings.append(f"Row {row}: Missing unit description")

        if contracted_units in [None, "", 0]:
            validation_warnings.append(f"Row {row}: Missing or zero contracted units")

        if unit_cost in [None, "", 0]:
            validation_warnings.append(f"Row {row}: Missing or zero unit cost")

        if forecast_units and contracted_units and abs(forecast_units - contracted_units) > 0.01:
            validation_warnings.append(
                f"Row {row}: Forecast units do not match contracted units"
            )

        if contracted_units and unit_cost:
            total_contract_value += contracted_units * unit_cost

        if forecast_units:
            total_forecast_units += forecast_units

    if missing_row_timeline_count > 0:
        validation_warnings.append(
            f"{missing_row_timeline_count} rows had missing activity-level dates and used project-level timeline fallback."
        )

    high_risk = []
    medium_risk = []
    low_risk = []

    for warning in validation_warnings:
        text = warning.lower()

        if "missing or zero unit cost" in text:
            high_risk.append(warning)
        elif "forecast units do not match contracted units" in text:
            medium_risk.append(warning)
        elif "missing activity name" in text:
            medium_risk.append(warning)
        elif "missing unit description" in text:
            medium_risk.append(warning)
        elif "missing or zero contracted units" in text:
            medium_risk.append(warning)
        elif "fallback" in text:
            medium_risk.append(warning)
        else:
            low_risk.append(warning)

    if len(high_risk) == 0 and len(medium_risk) <= 2:
        confidence = "High"
    elif len(high_risk) <= 2:
        confidence = "Medium"
    else:
        confidence = "Low"

    actions = []

    if high_risk:
        actions.append("Review revenue-impacting rows immediately.")

    if medium_risk:
        actions.append("Validate warnings, especially missing row-level timelines.")

    if total_forecast_units > 0:
        actions.append("Confirm forecast spread aligns with row-level activity timelines.")

    if total_contract_value > 0:
        actions.append("Verify contract totals against budget assumptions.")

    if "AI PMO Review" in wb.sheetnames:
        del wb["AI PMO Review"]

    review_ws = wb.create_sheet("AI PMO Review")

    review_ws["A1"] = "P95 AI PMO REVIEW"
    review_ws["A3"] = "Revenue Confidence"
    review_ws["B3"] = confidence
    review_ws["A4"] = "Rows Processed"
    review_ws["B4"] = rows_processed
    review_ws["A5"] = "Warnings"
    review_ws["B5"] = len(validation_warnings)
    review_ws["A6"] = "High Risk"
    review_ws["B6"] = len(high_risk)
    review_ws["A7"] = "Medium Risk"
    review_ws["B7"] = len(medium_risk)
    review_ws["A8"] = "Low Risk"
    review_ws["B8"] = len(low_risk)
    review_ws["A9"] = "Total Contract Value"
    review_ws["B9"] = round(total_contract_value, 2)
    review_ws["A10"] = "Total Forecast Units"
    review_ws["B10"] = round(total_forecast_units, 2)
    review_ws["A11"] = "Budget Activity Source Sheet"
    review_ws["B11"] = activity_sheet
    review_ws["A12"] = "Project Timeline Source Sheet"
    review_ws["B12"] = timeline_source["sheet"]
    review_ws["A13"] = "Project Timeline Source Row"
    review_ws["B13"] = timeline_source["row"]
    review_ws["A14"] = "Detected Project Start"
    review_ws["B14"] = project_start
    review_ws["A15"] = "Detected Project End"
    review_ws["B15"] = project_end
    review_ws["A16"] = "Rows Missing Activity-Level Timeline"
    review_ws["B16"] = missing_row_timeline_count

    row_num = 18

    review_ws[f"A{row_num}"] = "HIGH-RISK ITEMS"
    row_num += 1
    for item in high_risk:
        review_ws[f"A{row_num}"] = item
        row_num += 1

    row_num += 1
    review_ws[f"A{row_num}"] = "MEDIUM-RISK ITEMS"
    row_num += 1
    for item in medium_risk:
        review_ws[f"A{row_num}"] = item
        row_num += 1

    row_num += 1
    review_ws[f"A{row_num}"] = "LOW-RISK ITEMS"
    row_num += 1
    for item in low_risk:
        review_ws[f"A{row_num}"] = item
        row_num += 1

    row_num += 2
    review_ws[f"A{row_num}"] = "SUGGESTED PMO ACTIONS"
    row_num += 1
    for action in actions:
        review_ws[f"A{row_num}"] = action
        row_num += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    summary = {
        "confidence": confidence,
        "rows_processed": rows_processed,
        "warnings": len(validation_warnings),
        "high_risk": len(high_risk),
        "medium_risk": len(medium_risk),
        "low_risk": len(low_risk),
        "total_contract_value": round(total_contract_value, 2),
        "total_forecast_units": round(total_forecast_units, 2),
        "actions": actions,
        "activity_sheet": activity_sheet,
        "timeline_source": timeline_source,
        "project_start": project_start,
        "project_end": project_end,
        "missing_row_timeline_count": missing_row_timeline_count
    }

    return output, summary


st.sidebar.header("Study Parameters")
st.sidebar.info(
    "Budget lines are read from 'Workload and Resources P1'. "
    "Activity-level timelines are taken from start time/end time columns. "
    "If missing, the project-level timeline is used as fallback."
)

col1, col2 = st.columns(2)

with col1:
    st.subheader("Budget File")
    budget_file = st.file_uploader("Upload Budget Excel (.xlsx)", type=["xlsx"])

with col2:
    st.subheader("Unit Tracker Template")
    template_file = st.file_uploader("Upload Unit Tracker Template (.xlsx)", type=["xlsx"])

if st.button("Generate Revenue Module"):
    if not budget_file or not template_file:
        st.error("Please upload both files.")
    else:
        try:
            output, summary = generate_revenue_tracker(budget_file, template_file)

            st.success("Revenue module generated successfully.")

            st.subheader("Detected Sources")
            st.write(f"Budget activity data used from sheet: **{summary['activity_sheet']}**")
            st.write(
                f"Project timeline detected from sheet: **{summary['timeline_source']['sheet']}**, "
                f"row **{summary['timeline_source']['row']}**"
            )
            st.write(f"Detected project start: **{summary['project_start'].strftime('%d-%b-%Y')}**")
            st.write(f"Detected project end: **{summary['project_end'].strftime('%d-%b-%Y')}**")
            st.write(
                f"Rows using project-level fallback timeline: "
                f"**{summary['missing_row_timeline_count']}**"
            )

            st.subheader("PMO Review Summary")
            st.write(f"Revenue Confidence: **{summary['confidence']}**")
            st.write(f"Rows Processed: **{summary['rows_processed']}**")
            st.write(f"Warnings: **{summary['warnings']}**")
            st.write(f"Total Contract Value: **{summary['total_contract_value']}**")
            st.write(f"Total Forecast Units: **{summary['total_forecast_units']}**")

            st.subheader("Suggested PMO Actions")
            for action in summary["actions"]:
                st.write(f"- {action}")

            st.download_button(
                label="Download Completed Revenue Tracker",
                data=output,
                file_name="P95_UNIT_TRACKER_ROW_TIMELINES.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error("Something went wrong while generating the tracker.")
            st.exception(e)

st.divider()
st.caption("P95 AI Revenue Module Generator • Powered by Dania Alqudah")
```

Commit and refresh.
